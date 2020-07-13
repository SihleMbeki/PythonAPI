import openpyxl
import pytest
import requests
import json
import jsonpath

@pytest.fixture()
def setFile():
    global file
    global wb
    global limit
    limit=100
    wb = openpyxl.Workbook()
    file ="C:\\Users\\Sihle\\Documents\\Career Documentation\\demo.xlsx"

    yield
    # Anytime you modify the Workbook object
    # or its sheets and cells, the spreadsheet
    # file will not be saved until you call
    # the save() workbook method.
    wb.save(file)

def test_Interests(setFile):
    params={'type':'adinterest','q':'[Golf]','limit':limit+1,'locale':'en_US','access_token':'652146028717265|sJpbfMjhDvZiwYGsXYsf_MQrOtQ'}
    url='https://graph.facebook.com/search?'
    response=requests.get(url,params=params)
    assert response.status_code ==200

    # Get workbook active sheet
    # from the active attribute
    sheet = wb.active

    # Cell objects also have row, column
    # and coordinate attributes that provide
    # location information for the cell.

    # Note: The first row or column integer
    # is 1, not 0. Cell object is created by
    # using sheet object's cell() method.
    response_json = json.loads(response.text)
    # A1 means column = 1 & row = 1.
    a1 = sheet['A1']
    a1.value = "Name"

    # B1 means column = 2 & row = 1.
    b1 = sheet['B1']
    b1.value = "Description"

    #C1 means column = 3 & row = 1.
    c1 = sheet['C1']
    c1.value = "Path"

    # D1 means column = 4 & row = 1.
    d1 = sheet['D1']
    d1.value = "Topic"
   # global row
    row=2
    for i in range(0,limit):
        topic = jsonpath.jsonpath(response_json, 'data['+str(i)+'].topic')
        name = jsonpath.jsonpath(response_json, 'data['+str(i)+'].name')
        path = jsonpath.jsonpath(response_json, 'data['+str(i)+'].path')
        description = jsonpath.jsonpath(response_json, 'data['+str(i)+'].description')
        #set column values for the current row
        try:
            col1 = sheet.cell(row=row, column=1)
            col1.value = name[0]
            col2 = sheet.cell(row=row, column=2)
            col2.value = str(description[0])
            col3 = sheet.cell(row=row, column=3)
            col3.value = str(path[0])
            col4 = sheet.cell(row=row, column=4)
            col4.value = str(topic[0])
            row+=1
        except:
            col1 = sheet.cell(row=row, column=1)
            col1.value='Error - object is not subscriptable'+str(row)
            row += 1





